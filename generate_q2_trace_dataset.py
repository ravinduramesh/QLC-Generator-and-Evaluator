from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl.reader.excel import load_workbook

DEFAULT_INPUT = "Lab9Responses.xlsx"
DEFAULT_OUTPUT = "Lab9Responses_Q2_traces-ALL.csv"


def _strip_markdown_code_fences(text: str) -> str:
    stripped = text.strip()
    if not stripped.startswith("```"):
        return stripped

    lines = stripped.splitlines()
    if not lines:
        return stripped

    if lines[0].startswith("```"):
        lines = lines[1:]
    if lines and lines[-1].strip() == "```":
        lines = lines[:-1]
    return "\n".join(lines).strip()


def _extract_json_object(text: str) -> str:
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end < start:
        return text
    return text[start:end + 1]


def _escape_unescaped_control_chars_in_json_strings(text: str) -> str:
    result: list[str] = []
    in_string = False
    escaped = False

    for ch in text:
        if not in_string:
            result.append(ch)
            if ch == '"':
                in_string = True
            continue

        if escaped:
            result.append(ch)
            escaped = False
            continue

        if ch == "\\":
            result.append(ch)
            escaped = True
            continue

        if ch == '"':
            result.append(ch)
            in_string = False
            continue

        codepoint = ord(ch)
        if codepoint < 0x20:
            if ch == "\n":
                result.append("\\n")
            elif ch == "\r":
                result.append("\\r")
            elif ch == "\t":
                result.append("\\t")
            else:
                result.append(f"\\u{codepoint:04x}")
            continue

        result.append(ch)

    return "".join(result)


def parse_model_json(raw: str) -> dict[str, Any]:
    candidates = [
        raw,
        _strip_markdown_code_fences(raw),
        _extract_json_object(raw),
        _escape_unescaped_control_chars_in_json_strings(raw),
        _escape_unescaped_control_chars_in_json_strings(_extract_json_object(_strip_markdown_code_fences(raw))),
    ]

    last_error: Exception | None = None
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict):
                return parsed
        except Exception as exc:  # noqa: BLE001 - keep trying best-effort parse variants
            last_error = exc

    if last_error is None:
        raise ValueError("Model response could not be parsed as a JSON object")
    raise ValueError("Model response could not be parsed as a JSON object") from last_error


def clean_excel_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("_x000D_\n", "\n")
    text = text.replace("_x000D_", "\n")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip("\ufeff")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read the Lab9Responses workbook, send each Response 2 code submission "
            "to OpenAI, and write questions plus modified code variants to CSV."
        )
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
        help=f"Path to the source xlsx file (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Path to the output CSV file (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional worksheet name. If omitted, the first sheet is used.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Optional limit for the number of student submissions to process.",
    )
    return parser.parse_args()


def get_response_2_column(headers: list[str]) -> int:
    for index, header in enumerate(headers, start=1):
        if header.strip().lower() == "response 2":
            return index
    return 7


def build_prompt(student_code: str) -> list[dict[str, str]]:
    system_prompt = (
        "You generate educational code-tracing material for student C submissions. "
        "Return only valid JSON with the exact keys question_1, question_2, answer_1, answer_2, "
        "modified_code_1, and modified_code_2. Each value must be a string."
    )
    user_prompt = f"""
Create two distinct code-tracing questions for this student submission, then provide two modified code variants.
Also generate a concise answer for each question.
Code tracing questions should not be based on the final output. Ask questions to trace some parts of the code to check students understanding about the parts of the code.

Student code:
```c
{student_code}
```

Rules:
- Keep the student's code body verbatim.
- Only add lines that are necessary to run a hard-coded test and print the correct answer.
- The added lines should not alter the student's implementation logic.
- The modified code should be a standalone C snippet if the original code already contains only the function definition.
- The two questions should be different and directly based on tracing the provided code.
- The two answers should directly answer their corresponding questions.
- Make the question wording concise and classroom-friendly.
- Make the answers concise and classroom-friendly.
- The modified code must print only the answer for that question.
- Do not include markdown fences in the JSON values.
""".strip()
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def call_openai_for_submission(client: OpenAI, model: str, student_code: str) -> dict[str, str]:
    required_keys = ["question_1", "question_2", "answer_1", "answer_2", "modified_code_1", "modified_code_2"]

    for attempt in range(1, 4):
        response = client.chat.completions.create(
            model=model,
            messages=build_prompt(student_code)
        )
        raw = response.choices[0].message.content or "{}"

        try:
            payload = parse_model_json(raw)
            missing = [key for key in required_keys if key not in payload]
            if missing:
                raise ValueError(f"OpenAI response missing keys: {', '.join(missing)}")
            return {key: str(payload[key]) for key in required_keys}
        except Exception as exc:  # noqa: BLE001 - retry on malformed model output
            print(
                f"Warning: malformed model response on attempt {attempt}/3: {exc}",
                file=sys.stderr,
            )
            if attempt == 3:
                snippet = raw[:500].replace("\n", "\\n")
                raise ValueError(
                    "Failed to parse model response after 3 attempts. "
                    f"Response snippet: {snippet}"
                ) from exc

    raise RuntimeError("Unreachable")


def read_submissions(workbook_path: Path, sheet_name: str | None) -> tuple[list[str], list[str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    headers = [clean_excel_text(worksheet.cell(1, column).value) for column in range(1, worksheet.max_column + 1)]
    response_2_column = get_response_2_column(headers)

    submissions: list[str] = []
    for row in range(2, worksheet.max_row + 1):
        code = clean_excel_text(worksheet.cell(row, response_2_column).value)
        if code:
            submissions.append(code)

    return headers, submissions


def process_workbook(input_path: Path, output_path: Path, sheet_name: str | None, limit: int | None) -> None:
    load_dotenv()
    client = OpenAI(
        base_url=os.getenv("QA_GENERATOR_BASE_URL"),
        api_key=os.getenv("QA_GENERATOR_API_KEY")
    )
    model = os.getenv("QA_GENERATOR_MODEL")

    _, submissions = read_submissions(input_path, sheet_name)

    if limit is not None:
        submissions = submissions[:limit]

    fieldnames = [
        "student_code",
        "question_1",
        "question_2",
        "answer_1",
        "answer_2",
        "modified_code_1",
        "modified_code_2",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()

        for index, student_code in enumerate(submissions, start=1):
            print(f"Processing submission {index}/{len(submissions)}...", file=sys.stderr)
            generated = call_openai_for_submission(client, model, student_code)
            writer.writerow(
                {
                    "student_code": student_code,
                    "question_1": generated["question_1"],
                    "question_2": generated["question_2"],
                    "answer_1": generated["answer_1"],
                    "answer_2": generated["answer_2"],
                    "modified_code_1": generated["modified_code_1"],
                    "modified_code_2": generated["modified_code_2"],
                }
            )


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise SystemExit(f"Input workbook not found: {input_path}")

    process_workbook(input_path, output_path, args.sheet, args.limit)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
