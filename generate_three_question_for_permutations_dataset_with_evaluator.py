from __future__ import annotations

import csv
import itertools
import json
import os
import sys
import unicodedata
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl.reader.excel import load_workbook

DEFAULT_INPUT = "Lab9Responses-Mini.xlsx"
DEFAULT_OUTPUT = "Lab9Responses_three_questions_permutations.csv"
DEFAULT_SHEET = None
DEFAULT_LIMIT = None

TARGET_LEVELS = [
    (
        "Understand",
        "Explain the purpose or intent of syntax, keywords, or language rules without executing it manually line by line. (ex: What does the remove() method do?, What is the purpose of i++?, what is the purpose of using 'length - 1 - i' as an index for the word array?)",
    ),
    (
        "Apply",
        "Apply operator precedence rules, solving logic and iteration counts, or tracing code with specified inputs to determine the exact state or output.",
    ),
    (
        "Create",
        "Suggest necessary changes to make the code style better, fix a bug, make the code more memory efficient, make the code more runtime efficient. If the code is perfect, modify the code to support a new test case (a minor new feature)",
    ),
]

LEVEL_PERMUTATIONS = list(itertools.permutations([level for level, _ in TARGET_LEVELS]))
LEVEL_DESCRIPTIONS = {level: description for level, description in TARGET_LEVELS}

EVALUATION_FIELDS = {
    "Correctness": bool,
    "AnswerCorrectness": bool,
    "AnswerCompleteness": bool,
    "Grammatical": bool,
    "Clarity": bool,
    "RelevanceToCourseContent": bool,
    "RelevanceToLearnerCode": int,
    "AppropriatenessOfDifficultyForCS1": int,
    "RevisedBloomsTaxonomyLevel": str,
}
VALID_BLOOM_LEVELS = {"Remember", "Understand", "Apply", "Analyse", "Evaluate", "Create"}


def _extract_json_object(text: str) -> str:
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end < start:
        return text
    return text[start : end + 1]


def _escape_unescaped_control_chars_in_json_strings(text: str) -> str:
    result: list[str] = []
    in_string = False
    escaped = False

    for ch in text:
        if not in_string:
            result.append(ch)
            if ch == '"':
                in_string = True
            continue

        if escaped:
            result.append(ch)
            escaped = False
            continue

        if ch == "\\":
            result.append(ch)
            escaped = True
            continue

        if ch == '"':
            result.append(ch)
            in_string = False
            continue

        codepoint = ord(ch)
        if codepoint < 0x20:
            if ch == "\n":
                result.append("\\n")
            elif ch == "\r":
                result.append("\\r")
            elif ch == "\t":
                result.append("\\t")
            else:
                result.append(f"\\u{codepoint:04x}")
            continue

        result.append(ch)

    return "".join(result)


def parse_model_json(raw: str) -> dict[str, Any]:
    candidates = [
        raw,
        _extract_json_object(raw),
        _escape_unescaped_control_chars_in_json_strings(raw),
    ]

    last_error: Exception | None = None
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict):
                return parsed
        except Exception as exc:  # noqa: BLE001 - keep trying best-effort parse variants
            last_error = exc

    if last_error is None:
        raise ValueError("Model response could not be parsed as a JSON object")
    raise ValueError("Model response could not be parsed as a JSON object") from last_error


def normalize_text(s: str) -> str:
    return unicodedata.normalize("NFKC", s)


def clean_excel_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("_x000D_\n", "\n")
    text = text.replace("_x000D_", "\n")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = normalize_text(text.strip("\ufeff"))
    return text


def get_column_index(headers: list[str], candidates: list[str]) -> int | None:
    normalized_headers = [header.strip().lower() for header in headers]
    for candidate in candidates:
        normalized_candidate = candidate.strip().lower()
        for idx, header in enumerate(normalized_headers):
            if header == normalized_candidate:
                return idx
    return None


def get_response_2_column(headers: list[str]) -> int:
    for index, header in enumerate(headers):
        if header.strip().lower() == "response 2":
            return index
    return 6


def build_question_generation_prompt(student_code: str, target_level: str, level_description: str, problem_statement: str = "A palindrome is a word that reads exactly the same from left to right or from right to left (an example is “noon”).  Write a function called IsPalindrome() which takes a string as input, and returns 1 (i.e. true) if that string is a palindrome and 0 (i.e. false) otherwise.  You can assume that all characters in the string will be in lower case, and the input string will contain at least one character. NOTE: You can assume that the <string.h> header file is included, and therefore you can use the strlen() function") -> list[dict[str, str]]:
    system_prompt = ("""
Generate a concise, introductory programming classroom-friendly short-answer question that deeply probes a student's understanding of their own code with respect to a given Bloom's taxonomy level and corresponding description, targeting a specific aspect of the code (e.g., a particular line, construct, or logical element).  
Accept the following required inputs:
- Student's code (as a string; do not modify or execute).
- The problem statement provided to the student (as context).
- A single Bloom's taxonomy level.
- The description for the specific Bloom's taxonomy level.

**Task details:**
1. Carefully and silently review the student's code in the context of the problem statement.
2. Silently identify a critical, conceptually important, or potentially misunderstood element or construct in the code that offers a strong opportunity to probe the specified Bloom's taxonomy level per its description.
3. Silently ensure that the question is concise, precise, introductory programming classroom-friendly short-answer question targeting the given Bloom’s level and specifically pointing out a part of the given code.
4. Silently ensure the question does not require code-writing, nor line-by-line code trace as the answer, nor depend on the code's final output.
5. Silently ensure the resulting question has a single correct answer and question specifically ask for that answer.
6. Output only valid JSON with exactly these keys (as plain text values):
   - "question": [question as one or two sentences. only single question. don't ask two questions]
   - "answer": [one or two sentences; the correct answer to the question]

**Formatting requirements:**  
- Output pure JSON only, using only the keys "question" and "answer".
- Do not include explanations, code snippets, or any other output.

**Example input:**  
- code: `for i in range(len(lst)): lst[i] = lst[i]*2`
- problem_statement: "Write code that doubles every element in a list."
- bloom_level: "Understand"
- bloom_desc: "explain the purpose or intent of syntax, keywords, or language rules without executing it manually line by line."

**Example output:**
{
  "question": "What is the role of 'lst[i] = lst[i]*2' in this code?",
  "answer": "It updates the element in the index i of the list by multiplying its value by two."
}

**Edge cases to consider:**
- If multiple meaningful target areas exist, prioritize a conceptually important or commonly misunderstood point.
- If the code is trivial or the context is unclear, ask about the most unique or essential construct in the code.

---

**Important reminder:**  
- Only output valid JSON with "question" and "answer" fields, as described above.
- The question must never explicitly reference the final output of the code or require code-writing, but must probe understanding of the code’s mechanics or reasoning as tailored to the specified Bloom's level.
""")
    user_prompt = f"""
code:
```c
{student_code}
```

problem_statement: '{problem_statement}'

bloom_level: '{target_level}'
bloom_desc: '{level_description}'
""".strip()
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

def call_openai_for_question_generation(client: OpenAI, model: str, student_code: str, target_level: str, level_description: str) -> dict[str, str]:
    required_keys = ["question", "answer"]

    for attempt in range(1, 4):
        response = client.chat.completions.create(
            model=model,
            messages=build_question_generation_prompt(student_code, target_level, level_description),
        )
        raw = response.choices[0].message.content or "{}"

        try:
            payload = parse_model_json(raw)
            missing = [key for key in required_keys if key not in payload]
            if missing:
                raise ValueError(f"OpenAI response missing keys: {', '.join(missing)}")
            return {key: str(payload[key]) for key in required_keys}
        except Exception as exc:  # noqa: BLE001 - retry on malformed model output
            print(
                f"Warning: malformed model response on attempt {attempt}/3: {exc}",
                file=sys.stderr,
            )
            if attempt == 3:
                snippet = raw[:500].replace("\n", "\\n")
                raise ValueError(
                    "Failed to parse model response after 3 attempts. "
                    f"Response snippet: {snippet}"
                ) from exc

    raise RuntimeError("Unreachable")


def build_question_evaluation_prompt(
    student_code: str,
    question: str,
    answer: str,
) -> list[dict[str, str]]:
    system_prompt = """
Evaluate a question and answer pair generated about a student's code using the provided rubric. Carefully read the student's source code, the generated question, and its answer. For each rubric criterion below, provide a judgment. Before producing your final evaluation, explicitly reason through and justify each decision, considering any nuances, ambiguities, or edge cases. After justifying your reasoning for each, provide your findings as a structured JSON object.

If any rubric field could be interpreted ambiguously (e.g., revised Bloom's taxonomy), reflect on possible options and explain your logic before reaching a final conclusion. All individual field conclusions and ratings must come after their associated reasoning.

Continue evaluating and justifying each rubric field until all are completed. Only after all reasoning, supply your JSON output at the very end.

Rubric fields to evaluate:
- Correctness: Is the question factually accurate, complete, and free of falsehoods? (Boolean)
- Answer's Correctness: Is the answer to the question factually correct? (Boolean)
- Answer's Completeness: Does the answer fully address the question, especially if multiple possible aspects exist? (Boolean)
- Grammatical: Is the question grammatically correct? (Boolean)
- Clarity: Is the question clear and unambiguous? (Boolean)
- Relevance to course content: Can the question be answered solely from course content? (Boolean)
- Relevance to learner code: Does the question engage directly with aspects of the provided code (syntax, semantics, bugs, etc.)? (1-5 scale, with 5 = maximal relevance)
- Appropriateness of difficulty for CS1: Is the question suitable for an introductory programming course (CS1)? (1-5 scale, with 3 as average difficulty)
- Revised Bloom's Taxonomy Level: What cognitive process does the question primarily assess? (Remember, Understand, Apply, Analyse, Evaluate, Create — string)

Respond only with your evaluation and the required JSON object.

**Output format:**
- Detailed reasoning and justification for each rubric point (as full sentences or bullet points), in order, before any conclusions.
- Output JSON as the very last item in your answer. The JSON should include all rubric fields as keys, with your concluded value for each. Do not wrap the JSON in code blocks.

**Example**

[Begin Example]

**Reasoning:**
- Correctness: The question correctly describes the function behavior as shown in the code. There are no factual errors.
- Answer's Correctness: The answer accurately describes the output for the given code input.
- Answer's Completeness: There could be another edge case the answer does not mention, but since the question only asks about a single case, the answer covers all needed details.
- Grammatical: The question follows standard grammar and punctuation.
- Clarity: The phrasing is specific and leaves no ambiguity.
- Relevance to course content: Lists and loops are core CS1 topics.
- Relevance to learner code: The question asks directly about the logic used in the student's code, so this is highly relevant (5/5).
- Appropriateness of difficulty for CS1: The question asks about control flow, which is typical of CS1, and not unnecessarily difficult (3/5).
- Revised Bloom's Taxonomy Level: The question asks for an explanation, so this matches the "Understand" level.

**Evaluation JSON:**
{
  "Correctness": true,
  "AnswerCorrectness": true,
  "AnswerCompleteness": true,
  "Grammatical": true,
  "Clarity": true,
  "RelevanceToCourseContent": true,
  "RelevanceToLearnerCode": 5,
  "AppropriatenessOfDifficultyForCS1": 3,
  "RevisedBloomsTaxonomyLevel": "Understand"
}

[End Example]

Please repeat this evaluation process for every question/answer/code set you review. Remember to:
- Reason through all rubric fields first, in order.
- Only output your completed JSON after all reasoning steps.

**Important:**  
- Always justify each field's rating before giving your conclusion for that field.
- The final answer is a single JSON and includes one field for each rubric criterion.  
- Never present conclusions before their associated reasoning.

**Reminder:**  
Evaluate a programming question/answer pair about code using all rubric fields. Each field’s justification comes before its result. Yield a single JSON evaluation at the end.""".strip()
    user_prompt = f"""
student_code:
```c
{student_code}
```
question: {question}
answer: {answer}
""".strip()
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def validate_question_evaluation(payload: dict[str, Any]) -> dict[str, Any]:
    missing = [key for key in EVALUATION_FIELDS if key not in payload]
    if missing:
        raise ValueError(f"Evaluator response missing keys: {', '.join(missing)}")

    for key, expected_type in EVALUATION_FIELDS.items():
        value = payload[key]
        if expected_type is bool and type(value) is not bool:
            raise ValueError(f"Evaluator field '{key}' must be boolean")
        if expected_type is int and (type(value) is not int or not 1 <= value <= 5):
            raise ValueError(f"Evaluator field '{key}' must be an integer from 1 to 5")
        if expected_type is str and (not isinstance(value, str) or value not in VALID_BLOOM_LEVELS):
            raise ValueError(f"Evaluator field '{key}' must be a valid revised Bloom level")

    return {key: payload[key] for key in EVALUATION_FIELDS}


def call_evaluator_agent(
    client: OpenAI,
    model: str,
    student_code: str,
    question: str,
    answer: str,
) -> dict[str, Any]:
    for attempt in range(1, 4):
        response = client.chat.completions.create(
            model=model,
            messages=build_question_evaluation_prompt(
                student_code,
                question,
                answer,
            ),
        )
        raw = response.choices[0].message.content or "{}"
        try:
            return validate_question_evaluation(parse_model_json(raw))
        except Exception as exc:  # noqa: BLE001 - retry on malformed evaluator output
            print(
                f"Warning: malformed evaluator response on attempt {attempt}/3: {exc}",
                file=sys.stderr,
            )
            if attempt == 3:
                snippet = raw[:500].replace("\n", "\\n")
                raise ValueError(
                    "Failed to parse evaluator response after 3 attempts. "
                    f"Response snippet: {snippet}"
                ) from exc

    raise RuntimeError("Unreachable")


def read_submissions(workbook_path: Path, sheet_name: str | None) -> tuple[list[str], list[str], list[str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    headers = [clean_excel_text(worksheet.cell(1, column).value) for column in range(1, worksheet.max_column + 1)]

    anon_id_column = get_column_index(headers, ["ANON_ID", "anon_id", "anon id", "student id"])
    response_2_column = get_response_2_column(headers)

    anon_ids: list[str] = []
    submissions: list[str] = []
    for row in range(2, worksheet.max_row + 1):
        anon_id = ""
        if anon_id_column is not None:
            anon_id = clean_excel_text(worksheet.cell(row, anon_id_column + 1).value)
        code = clean_excel_text(worksheet.cell(row, response_2_column + 1).value)
        if code:
            anon_ids.append(anon_id)
            submissions.append(code)

    return headers, anon_ids, submissions


def generate_question_set(
    client: OpenAI,
    model: str,
    student_code: str,
    level_permutation: tuple[str, ...],
) -> list[dict[str, str]]:
    generated = []
    for target_level in level_permutation:
        generated.append(
            call_openai_for_question_generation(
                client,
                model,
                student_code,
                target_level,
                LEVEL_DESCRIPTIONS[target_level],
            )
        )
    return generated


def build_group_assignments(total_rows: int) -> list[int]:
    if total_rows <= 0:
        return []

    group_size = total_rows // 6
    assignments: list[int] = []
    for group_index in range(5):
        assignments.extend([group_index] * group_size)

    remaining = total_rows - (5 * group_size)
    assignments.extend([5] * remaining)
    return assignments


def process_workbook(input_path: Path, output_path: Path, sheet_name: str | None, limit: int | None) -> None:
    load_dotenv()
    client = OpenAI(
        base_url=os.getenv("QA_GENERATOR_BASE_URL"),
        api_key=os.getenv("QA_GENERATOR_API_KEY"),
    )
    model = os.getenv("QA_GENERATOR_MODEL")
    evaluator_client = OpenAI(
        base_url=os.getenv("EVALUATOR_BASE_URL"),
        api_key=os.getenv("EVALUATOR_API_KEY"),
    )
    evaluator_model = os.getenv("EVALUATOR_MODEL")

    _, anon_ids, submissions = read_submissions(input_path, sheet_name)

    if limit is not None:
        anon_ids = anon_ids[:limit]
        submissions = submissions[:limit]

    fieldnames = [
        "ANON_ID",
        "original_code",
        "q1",
        "a1",
        "q1_level",
        "q1_evaluation",
        "q2",
        "a2",
        "q2_level",
        "q2_evaluation",
        "q3",
        "a3",
        "q3_level",
        "q3_evaluation",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()

        total_rows = len(submissions)
        group_assignments = build_group_assignments(total_rows)

        for index, (anon_id, student_code) in enumerate(zip(anon_ids, submissions), start=1):
            print(f"Processing submission {index}/{len(submissions)}...", file=sys.stderr)
            student_code = normalize_text(student_code)
            group_index = group_assignments[index - 1]
            level_permutation = LEVEL_PERMUTATIONS[group_index]
            generated_rows = generate_question_set(client, model, student_code, level_permutation)
            evaluations = [
                call_evaluator_agent(
                    evaluator_client,
                    evaluator_model,
                    student_code,
                    generated["question"],
                    generated["answer"],
                )
                for generated in generated_rows
            ]

            row = {
                "ANON_ID": normalize_text(anon_id),
                "original_code": student_code,
                "q1": normalize_text(generated_rows[0]["question"]),
                "a1": normalize_text(generated_rows[0]["answer"]),
                "q1_level": level_permutation[0],
                "q1_evaluation": json.dumps(evaluations[0], ensure_ascii=False),
                "q2": normalize_text(generated_rows[1]["question"]),
                "a2": normalize_text(generated_rows[1]["answer"]),
                "q2_level": level_permutation[1],
                "q2_evaluation": json.dumps(evaluations[1], ensure_ascii=False),
                "q3": normalize_text(generated_rows[2]["question"]),
                "a3": normalize_text(generated_rows[2]["answer"]),
                "q3_level": level_permutation[2],
                "q3_evaluation": json.dumps(evaluations[2], ensure_ascii=False),
            }
            writer.writerow(row)


def main() -> None:
    input_path = Path(DEFAULT_INPUT)
    output_path = Path(DEFAULT_OUTPUT)

    if not input_path.exists():
        raise SystemExit(f"Input workbook not found: {input_path}")

    process_workbook(input_path, output_path, DEFAULT_SHEET, DEFAULT_LIMIT)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()